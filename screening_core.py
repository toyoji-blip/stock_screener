"""
screening_core.py
==================
株式財務スクリーニングツールのコアロジック。
元の CLI 版 (screening_tool.py) から入出力（input/print）部分を取り除き、
Web アプリ（app.py）から呼び出せる純粋な関数群として再構成したもの。

【必要ライブラリ】
  pip install yfinance requests beautifulsoup4 lxml openpyxl matplotlib pytz flask
"""

import warnings
warnings.filterwarnings("ignore")

import json
import os
import shutil
import unicodedata
import uuid
from datetime import datetime, timedelta

import matplotlib
matplotlib.use("Agg")  # サーバー環境（GUIなし）用バックエンド

import yfinance as yf
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mticker
import pytz
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import logging
logging.getLogger("matplotlib.font_manager").setLevel(logging.ERROR)
matplotlib.rcParams["font.family"] = [
    "Meiryo", "Yu Gothic", "MS Gothic",
    "Noto Sans CJK JP", "IPAexGothic", "TakaoGothic", "sans-serif",
]
matplotlib.rcParams["axes.unicode_minus"] = False

TZ_JP = pytz.timezone("Asia/Tokyo")

# Excel/PNGの出力先（Webアプリ内の outputs フォルダ）
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 過去の分析結果（履歴）の保存先
HISTORY_PATH = os.path.join(OUTPUT_DIR, "history.json")
ARCHIVE_DIR = os.path.join(OUTPUT_DIR, "archive")
os.makedirs(ARCHIVE_DIR, exist_ok=True)


# ══════════════════════════════════════════════════════════════
#  ユーティリティ
# ══════════════════════════════════════════════════════════════

def nkc(s):
    return unicodedata.normalize("NFKC", str(s)).strip()


def to_m(v):
    """円 → 百万円"""
    return None if (v is None or pd.isna(v)) else round(float(v) / 1_000_000, 1)


def get_row(df, *kws):
    if df is None or df.empty:
        return None
    for kw in kws:
        kn = nkc(kw).lower()
        for idx in df.index:
            if kn in nkc(str(idx)).lower():
                return df.loc[idx]
    return None


def cagr(vals):
    """
    リストから CAGR（年平均成長率）を計算する。
    None を除いた最古→最新の2点で算出し、2期以上あれば計算する。

    CAGRは「開始値・終了値がともに正の値」である場合にのみ数学的に意味を持つ。
    どちらかが0以下（赤字・純有利子負債など）だと、例えば赤字が拡大している
    （-10→-40のように、より大きな赤字になっている）ケースでも
    比率としては (-40)/(-10)=4.0 と正の値になってしまい、
    見かけ上「+100%成長」のような誤った結果が出てしまう。
    そのため、開始値・終了値のいずれかが0以下の場合は None（計算不可）を返す。
    """
    # None を除いた有効値リスト
    v = [x for x in vals if x is not None]
    # 3期分のデータが揃っていない場合はCAGRを表示しない
    # （2期分だと「1年間の変化率」になってしまい、3年CAGRとして誤解を招く）
    if len(v) < len(vals):
        return None
    if len(v) < 2:
        return None
    first, last = v[0], v[-1]
    # 無配転落（first > 0 かつ last == 0）→ -100% を返す
    if first > 0 and last == 0:
        return -100.0
    # 開始値・終了値のいずれかが0以下の場合はCAGRが数学的に無意味
    if first <= 0 or last <= 0:
        return None
    n = len(v) - 1
    try:
        r = (last / first) ** (1 / n) - 1
        return round(r * 100, 1)
    except Exception:
        return None


def safe_filename(name):
    return "".join(c for c in name if c.isalnum() or c in "_") or "screening"


def build_metric_rows(fys, res, cash_vals=None, div_pay_vals=None):
    """
    財務指標テーブルの行を組み立てる（Step2プレビュー・Step4完了画面・
    過去の分析詳細・Excel出力のすべてで共通利用する）。

    cash_vals（現預金・百万円）/ div_pay_vals（配当支払総額・百万円）は
    短信からの手入力（Step3）が済むまで None のままでよい。
    その場合、それらに依存する行（FCF配当性向・ネットキャッシュ・
    グロス配当余裕・ネット配当余裕）は値が定まらないため None（画面上は「—」）になる。

    戻り値: [{"label":..., "values":[...], "cagr": float|None, "fmt": "{:,.2f}"}, ...]
    """
    n = len(fys)
    cash_vals = list(cash_vals) if cash_vals is not None else [None] * n
    div_pay_vals = list(div_pay_vals) if div_pay_vals is not None else [None] * n

    eps_vals = [res[fy].get("EPS") for fy in fys]
    dps_vals = [res[fy].get("DPS") for fy in fys]
    fcf_vals = [res[fy].get("FCF") for fy in fys]
    debt_vals = [res[fy].get("有利子負債") for fy in fys]

    # EPS配当性向[%] = DPS ÷ EPS × 100
    eps_payout = []
    for eps, dps in zip(eps_vals, dps_vals):
        if eps is not None and dps is not None and eps != 0:
            eps_payout.append(round(dps / eps * 100, 1))
        else:
            eps_payout.append(None)

    # FCF配当性向[%] = 配当支払総額 ÷ FCF × 100
    # 無配（dv=0）の場合は 0% と表示する。
    # FCF自体はdvが0でも表示されるべきなので、dv is not None かつ fcf is not None で判定する。
    fcf_payout = []
    for fcf, dv in zip(fcf_vals, div_pay_vals):
        if fcf is not None and dv is not None and fcf != 0:
            fcf_payout.append(round(dv / fcf * 100, 1))
        else:
            fcf_payout.append(None)

    # ネットキャッシュ（百万円） = 現預金 － 有利子負債
    net_cash = []
    for cash, debt in zip(cash_vals, debt_vals):
        if cash is not None and debt is not None:
            net_cash.append(round(cash - debt, 1))
        else:
            net_cash.append(None)

    # グロス配当余裕[年] = 現預金 ÷ 配当支払総額
    gross_headroom = []
    for cash, dv in zip(cash_vals, div_pay_vals):
        if cash is not None and dv is not None and dv != 0:
            gross_headroom.append(round(cash / dv, 1))
        else:
            gross_headroom.append(None)

    # ネット配当余裕[年] = ネットキャッシュ ÷ 配当支払総額
    # （有利子負債も考慮した、財務健全性ベースの余裕度。
    #   現預金は潤沢でも借入も多い会社では、グロスとネットで結果が
    #   大きく食い違うことがあるため、両方を併記して多面的に見られるようにする）
    net_headroom = []
    for nc, dv in zip(net_cash, div_pay_vals):
        if nc is not None and dv is not None and dv != 0:
            net_headroom.append(round(nc / dv, 1))
        else:
            net_headroom.append(None)

    return [
        {"label": "EPS（円）", "values": eps_vals, "cagr": cagr(eps_vals), "fmt": "{:,.2f}"},
        {"label": "DPS実績（円）", "values": dps_vals, "cagr": cagr(dps_vals), "fmt": "{:,.2f}"},
        {"label": "EPS配当性向[%]", "values": eps_payout, "cagr": None, "fmt": "{:,.1f}"},
        {"label": "FCF（百万円）", "values": fcf_vals, "cagr": cagr(fcf_vals), "fmt": "{:,.1f}"},
        {"label": "FCF配当性向[%]", "values": fcf_payout, "cagr": None, "fmt": "{:,.1f}"},
        {"label": "有利子負債（百万円）", "values": debt_vals, "cagr": cagr(debt_vals), "fmt": "{:,.1f}"},
        {"label": "現預金（百万円）", "values": cash_vals, "cagr": cagr(cash_vals), "fmt": "{:,.1f}"},
        {"label": "ネットキャッシュ（百万円）", "values": net_cash, "cagr": cagr(net_cash), "fmt": "{:,.1f}"},
        {"label": "グロス配当余裕[年]", "values": gross_headroom, "cagr": None, "fmt": "{:,.1f}"},
        {"label": "ネット配当余裕[年]", "values": net_headroom, "cagr": None, "fmt": "{:,.1f}"},
    ]


# ══════════════════════════════════════════════════════════════
#  過去の分析資料（履歴）
# ══════════════════════════════════════════════════════════════

def _load_history():
    if not os.path.exists(HISTORY_PATH):
        return []
    try:
        with open(HISTORY_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception:
        return []


def _save_history(entries):
    tmp_path = HISTORY_PATH + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, HISTORY_PATH)


def append_history_entry(entry):
    """1件分の分析結果を履歴に追記して保存する。"""
    entries = _load_history()
    entries.append(entry)
    _save_history(entries)
    return entry


def search_history(query=""):
    """
    銘柄コード／ティッカー／銘柄名の部分一致で履歴を検索する。
    query が空文字の場合は全件を新しい順で返す。
    """
    entries = _load_history()
    q = nkc(query or "").lower()
    if q:
        def hit(e):
            return (q in nkc(e.get("code", "")).lower()
                    or q in nkc(e.get("ticker", "")).lower()
                    or q in nkc(e.get("name", "")).lower())
        entries = [e for e in entries if hit(e)]
    entries.sort(key=lambda e: e.get("created_at", ""), reverse=True)
    return entries


def get_history_entry(entry_id):
    for e in _load_history():
        if e.get("id") == entry_id:
            return e
    return None


def delete_history_entry(entry_id):
    """
    履歴を1件削除する。history.json からのレコード削除に加えて、
    関連するExcelファイル（outputs/直下）とアーカイブ画像フォルダ
    （outputs/archive/<entry_id>/）も削除する。
    戻り値: 削除できた場合 True、該当データが無かった場合 False。
    """
    entries = _load_history()
    target = None
    remaining = []
    for e in entries:
        if e.get("id") == entry_id:
            target = e
        else:
            remaining.append(e)

    if target is None:
        return False

    _save_history(remaining)

    filename = target.get("filename")
    if filename:
        xlsx_path = os.path.join(OUTPUT_DIR, filename)
        if os.path.exists(xlsx_path):
            try:
                os.remove(xlsx_path)
            except Exception:
                pass

    archive_folder = os.path.join(ARCHIVE_DIR, entry_id)
    if os.path.isdir(archive_folder):
        try:
            shutil.rmtree(archive_folder)
        except Exception:
            pass

    return True


def save_archive_images(entry_id, fig_a, fig_b, fig_c):
    """
    グラフ画像を履歴閲覧用にアーカイブへ保存する。
    Excelに貼り付けた後も削除される一時PNGとは別に、ARCHIVE_DIR配下に
    entry_idごとのフォルダを作って保持する。
    戻り値: {"fig_a": "<entry_id>/fig_a.png", ...}（存在するものだけ）
    """
    folder = os.path.join(ARCHIVE_DIR, entry_id)
    os.makedirs(folder, exist_ok=True)
    rel_paths = {}
    for key, fig in [("fig_a", fig_a), ("fig_b", fig_b), ("fig_c", fig_c)]:
        if fig is not None:
            path = os.path.join(folder, f"{key}.png")
            fig.savefig(path, dpi=100, bbox_inches="tight")
            rel_paths[key] = f"{entry_id}/{key}.png"
    return rel_paths


# ══════════════════════════════════════════════════════════════
#  Step1: 銘柄基本情報取得
# ══════════════════════════════════════════════════════════════

def fetch_basic(code):
    """銘柄コードから基本情報（名称・株価・決算月推定）を取得する。"""
    tc = code if code.endswith(".T") else f"{code}.T"
    t = yf.Ticker(tc)
    info = t.info

    if not info or (info.get("longName") is None and info.get("shortName") is None
                     and info.get("currentPrice") is None and info.get("regularMarketPrice") is None):
        raise ValueError(f"銘柄コード '{code}' の情報が取得できませんでした。コードを確認してください。")

    name = info.get("longName") or info.get("shortName") or tc
    price = info.get("currentPrice") or info.get("regularMarketPrice")

    fy_end = info.get("fiscalYearEnd") or ""
    month_map = {"january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
                 "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12}
    fiscal_month = month_map.get(fy_end.lower()) if fy_end else None

    return {
        "ticker": tc,
        "name": name,
        "price": price,
        "fiscal_month_auto": fiscal_month,
    }


# ══════════════════════════════════════════════════════════════
#  Step2: 過去3年財務データ取得
# ══════════════════════════════════════════════════════════════

def fetch_financials(ticker, fiscal_month):
    """
    ticker: yf.Ticker 文字列（例: "4914.T"）または yf.Ticker インスタンス
    fiscal_month: 決算月（1〜12）
    戻り値: (fys:list[int], res:dict[int, dict])
    """
    t = ticker if hasattr(ticker, "info") else yf.Ticker(ticker)

    today = datetime.now(TZ_JP)
    latest_fy = today.year if today.month > fiscal_month else today.year - 1

    def fy_of(col):
        dt = pd.Timestamp(col)
        d = dt.tz_localize(TZ_JP) if dt.tzinfo is None else dt.tz_convert(TZ_JP)
        return d.year if d.month <= fiscal_month else d.year + 1

    # Yahoo Financeはfinancials種別（income/cash-flow/balance-sheet）ごとに
    # 独立して最大4期分を返す。そのため cash_flow が FY2021-2024 を返していても
    # balance_sheet は FY2022-2025 を返すことがある。
    # → 各種データソースの中で「最も新しいFY」を latest_fy として採用する。
    # → ただし未発表期（データが実在しない）は除外するため、
    #   いずれかのソースで確認できた最大FYを使う。
    is_ = None
    try:
        actual_fys = []

        # income_stmt から確認
        is_ = t.income_stmt
        if is_ is not None and not is_.empty:
            actual_fys += [fy_of(col) for col in is_.columns]

        # balance_sheet から確認（cash_flowより新しい期を持つことがある）
        bs_probe = t.balance_sheet
        if bs_probe is not None and not bs_probe.empty:
            actual_fys += [fy_of(col) for col in bs_probe.columns]

        # cash_flow から確認
        cf_probe = t.cash_flow
        if cf_probe is not None and not cf_probe.empty:
            actual_fys += [fy_of(col) for col in cf_probe.columns]

        if actual_fys:
            actual_latest = max(actual_fys)
            # 未発表期への行き過ぎを防ぐため、カレンダー推定値を上限とする
            if actual_latest < latest_fy:
                latest_fy = actual_latest
            elif actual_latest > latest_fy:
                latest_fy = actual_latest  # balance_sheetが新しい期を持っている場合に更新
    except Exception:
        pass

    fys = [latest_fy - 2, latest_fy - 1, latest_fy]
    res = {fy: {"EPS": None, "DPS": None, "FCF": None, "有利子負債": None} for fy in fys}

    info = t.info
    shares = info.get("sharesOutstanding") or info.get("impliedSharesOutstanding")

    # EPS
    try:
        if is_ is None:
            is_ = t.income_stmt
        ni = get_row(is_, "Net Income")
        if ni is not None and shares:
            for col in ni.index:
                fy = fy_of(col)
                if fy in res and pd.notna(ni[col]):
                    res[fy]["EPS"] = round(ni[col] / shares, 2)
    except Exception:
        pass

    # DPS（実績）
    try:
        divs = t.dividends
        if not divs.empty:
            for dt_, val in divs.items():
                d = dt_.tz_convert(TZ_JP) if dt_.tzinfo else dt_
                fy = d.year if d.month <= fiscal_month else d.year + 1
                if fy in res:
                    res[fy]["DPS"] = round((res[fy]["DPS"] or 0) + val, 2)
    except Exception:
        pass

    # FCF = 営業CF + 投資CF（Yahoo!ファイナンス日本式定義）
    # yfinanceが返す行名は銘柄・バージョンによって揺れるため、
    # 複数の候補キーワードで幅広くマッチさせる。
    try:
        cf = t.cash_flow

        # デバッグ用: 取得できた行名をすべてログ出力（本番でも残しておく）
        import logging
        _logger = logging.getLogger(__name__)
        _logger.debug("cash_flow rows: %s", list(cf.index) if cf is not None and not cf.empty else [])

        op = get_row(cf,
            "Operating Cash Flow",
            "Cash Flow From Continuing Operating Activities",
            "Total Cash From Operating Activities",
            "Net Cash Provided By Operating Activities",
            "Cash Generated From Operations",
            "operating")  # 部分一致で拾う

        inv = get_row(cf,
            "Investing Cash Flow",
            "Cash Flow From Continuing Investing Activities",
            "Total Cashflows From Investing Activities",
            "Net Cash Used In Investing Activities",
            "Net Cash Provided By Investing Activities",
            "investing")  # 部分一致で拾う

        if op is not None and inv is not None:
            # 本命: 営業CF + 投資CF
            for col in op.index:
                fy = fy_of(col)
                if (fy in res and pd.notna(op[col])
                        and col in inv.index and pd.notna(inv[col])):
                    res[fy]["FCF"] = to_m(op[col] + inv[col])

        if any(res[fy]["FCF"] is None for fy in fys):
            # フォールバック①: yfinance算出のFree Cash Flow行
            fcf_row = get_row(cf, "Free Cash Flow", "FreeCashFlow")
            if fcf_row is not None:
                for col in fcf_row.index:
                    fy = fy_of(col)
                    if fy in res and res[fy]["FCF"] is None and pd.notna(fcf_row[col]):
                        res[fy]["FCF"] = to_m(fcf_row[col])

        if any(res[fy]["FCF"] is None for fy in fys):
            # フォールバック②: 営業CF単体 + CapEx で近似
            if op is not None:
                cap = get_row(cf,
                    "Capital Expenditure", "Purchase Of PPE",
                    "Purchase of Property", "CapEx",
                    "capital expenditure")
                for col in op.index:
                    fy = fy_of(col)
                    if fy in res and res[fy]["FCF"] is None and pd.notna(op[col]):
                        c = cap[col] if (cap is not None and col in cap.index
                                          and pd.notna(cap[col])) else 0
                        res[fy]["FCF"] = to_m(op[col] + c)

        if any(res[fy]["FCF"] is None for fy in fys):
            # フォールバック③: 四半期データからFY分を合算
            # Yahoo Financeはyearly(年次)で最大4期しか返さないが、
            # quarterly(四半期)は直近5四半期を返すため、
            # 年次データに含まれない最新FYを四半期合算で補完できる。
            try:
                cf_q = t.get_cash_flow(freq="quarterly")
                if cf_q is not None and not cf_q.empty:
                    op_q  = get_row(cf_q, "Operating Cash Flow",
                                    "Cash Flow From Continuing Operating Activities",
                                    "Total Cash From Operating Activities",
                                    "Net Cash Provided By Operating Activities",
                                    "operating")
                    inv_q = get_row(cf_q, "Investing Cash Flow",
                                    "Cash Flow From Continuing Investing Activities",
                                    "Total Cashflows From Investing Activities",
                                    "Net Cash Used In Investing Activities",
                                    "investing")
                    if op_q is not None and inv_q is not None:
                        # 四半期ごとにFYを判定して合算
                        q_fcf = {}
                        for col in op_q.index:
                            fy = fy_of(col)
                            if fy in res and res[fy]["FCF"] is None:
                                op_v  = op_q[col]  if pd.notna(op_q[col])  else 0
                                inv_v = inv_q[col] if (col in inv_q.index and pd.notna(inv_q[col])) else 0
                                q_fcf[fy] = q_fcf.get(fy, 0) + op_v + inv_v
                        for fy, val in q_fcf.items():
                            if fy in res and res[fy]["FCF"] is None:
                                res[fy]["FCF"] = to_m(val)
            except Exception:
                pass
    except Exception:
        pass

    # 有利子負債
    try:
        bs = t.balance_sheet
        lt = get_row(bs, "Long Term Debt")
        st = get_row(bs, "Short Term Debt", "Current Debt", "Current Borrowings")
        src = lt if lt is not None else st
        if src is not None:
            for col in src.index:
                fy = fy_of(col)
                if fy in res:
                    lv = lt[col] if (lt is not None and col in lt.index
                                      and pd.notna(lt[col])) else 0
                    sv = st[col] if (st is not None and col in st.index
                                      and pd.notna(st[col])) else 0
                    res[fy]["有利子負債"] = to_m(lv + sv)
    except Exception:
        pass

    # yfinanceのデータ更新が遅れている場合、最新期のFCFがNoneになることがある
    # （例: 決算発表直後はyfinanceへの反映に数日〜数週間かかる場合がある）
    # その場合は画面上「—」と表示され、データが揃い次第自動的に取得される

    return fys, res


# ══════════════════════════════════════════════════════════════
#  グラフ描画
# ══════════════════════════════════════════════════════════════

COLORS = {
    "blue": "#4472C4", "red": "#C0504D", "green": "#9BBB59",
    "purple": "#8064A2", "orange": "#F79646", "gray": "#808080",
    "dkblue": "#17375E", "lblue": "#B8CCE4",
}


def make_fig_a(company, years, fcf_vals, div_vals, payout_vals):
    """
    グラフA: FCF・配当支払総額・FCF配当性向

    FCF配当性向の扱い:
      - FCF > 0: 右軸（0〜200%）に折れ線プロット
      - FCF < 0: 分母がマイナスのため折れ線から除外し、
                 X軸上（Y=0）に目立つアノテーションで実際の値を表示
    """
    fig, ax1 = plt.subplots(figsize=(7, 4.5))
    fig.patch.set_facecolor("#F2F2F2")
    ax1.set_facecolor("#FFFFFF")

    x = range(len(years))
    bw = 0.35
    ax1.bar([i - bw / 2 for i in x], fcf_vals, width=bw,
            label="FCF（百万円）", color=COLORS["blue"], zorder=3)
    ax1.bar([i + bw / 2 for i in x], div_vals, width=bw,
            label="配当支払総額（百万円）", color=COLORS["orange"], zorder=3)

    ax1.set_ylabel("金額（百万円）", fontsize=10)
    ax1.set_xticks(x)
    ax1.set_xticklabels(years, rotation=20)
    ax1.axhline(0, color="black", lw=0.8)
    ax1.grid(axis="y", ls="--", alpha=0.4, zorder=0)
    ax1.legend(loc="upper left", fontsize=9)

    ax2 = ax1.twinx()

    # FCFが正の期 → 折れ線プロット対象
    # FCFが負の期 → X軸上にアノテーション表示
    p_normal  = []   # 折れ線用（FCF>0かつpayout!=None）
    xi_normal = []
    neg_annots = []  # (index, raw_payout)

    for i, (fcf, raw) in enumerate(zip(fcf_vals, payout_vals)):
        if raw is None:
            p_normal.append(None)
            continue
        if fcf is not None and fcf < 0:
            neg_annots.append((i, raw))
            p_normal.append(None)
        else:
            p_normal.append(raw)
            xi_normal.append(i)

    # 右軸スケール
    valid_normal = [p for p in p_normal if p is not None]
    use_log = any(abs(p) > 200 for p in valid_normal)

    if use_log:
        p_plot = [max(abs(p), 0.1) if p is not None else None for p in p_normal]
        ax2.set_yscale("log")
        ax2.set_ylim(1, max([pp for pp in p_plot if pp] + [300]) * 2)
        ax2.set_ylabel("FCF配当性向（%）[対数]", fontsize=9, color=COLORS["green"])
    else:
        p_plot = p_normal
        ax2.set_ylim(0, 200)
        ax2.set_ylabel("FCF配当性向（%）", fontsize=9, color=COLORS["green"])

    # 折れ線: FCF>0 の期のみ
    xi_line = [i for i, p in enumerate(p_plot) if p is not None]
    yp_line = [p for p in p_plot if p is not None]
    if xi_line:
        ax2.plot(xi_line, yp_line, marker="o", color=COLORS["green"],
                 lw=2, label="FCF配当性向", zorder=5)
        for i, p in zip(xi_line, yp_line):
            raw = payout_vals[i]
            col = "crimson" if raw is not None and abs(raw) > 100 else "black"
            pct_label = "%.1f%%" % raw
            ax2.annotate(pct_label, (i, p),
                         xytext=(0, 8), textcoords="offset points",
                         fontsize=9, color=col, ha="center")

    # FCFがマイナスの期: X軸上（Y=0）に目立つアノテーション
    for idx, raw in neg_annots:
        pct_label = "%.1f%% (FCF<0)" % raw
        ax1.annotate(
            pct_label,
            xy=(idx, 0),
            xytext=(0, 6),
            textcoords="offset points",
            fontsize=10,
            fontweight="bold",
            color="crimson",
            ha="center",
            va="bottom",
            zorder=10,
            bbox=dict(boxstyle="round,pad=0.3", fc="lightyellow",
                      ec="crimson", lw=1.2, alpha=0.9),
        )

    ax2.legend(loc="upper right", fontsize=9)
    ax2.tick_params(axis="y", labelcolor=COLORS["green"])

    plt.title(company + "：FCF・配当支払総額・FCF配当性向", fontsize=11, pad=8)
    plt.tight_layout()
    return fig


def make_fig_b(company, years, cash_vals, debt_vals, div_vals):
    """グラフB: 現預金・有利子負債・配当支払総額"""
    fig, ax = plt.subplots(figsize=(7, 4.5))
    fig.patch.set_facecolor("#F2F2F2")
    ax.set_facecolor("#FFFFFF")

    x = range(len(years))
    bw = 0.25
    ax.bar([i for i in x], cash_vals, width=bw,
           label="現預金", color=COLORS["blue"], zorder=3)
    ax.bar([i + bw for i in x], div_vals, width=bw,
           label="配当支払総額", color=COLORS["green"], zorder=3)
    ax.bar([i + bw * 2 for i in x], debt_vals, width=bw,
           label="有利子負債", color=COLORS["purple"], zorder=3)

    ax.set_xticks([i + bw for i in x])
    ax.set_xticklabels(years, rotation=20)
    ax.set_ylabel("金額（百万円）", fontsize=10)
    ax.legend(fontsize=9)
    ax.grid(axis="y", ls="--", alpha=0.4, zorder=0)
    ax.axhline(0, color="black", lw=0.8)
    plt.title(f"{company}：現預金・有利子負債・配当支払総額", fontsize=11, pad=8)
    plt.tight_layout()
    return fig


def previous_business_day(now=None):
    """
    スクリーニング実行時点の「前日」を返す。
    前日が土日の場合は直近の金曜日にロールバックする。
    （例: 実行日が月曜 → 前日は日曜 → 金曜まで遡る／実行日が土日 → 金曜）
    """
    now = now or datetime.now(TZ_JP)
    d = now - timedelta(days=1)
    while d.weekday() >= 5:  # 5=土曜, 6=日曜
        d -= timedelta(days=1)
    return d


def make_fig_c(ticker, company, fiscal_month, forecast_dps, price, fys, res):
    """
    グラフC: 株価・配当利回りトレンド + 格安判定

    【実績配当利回りの計算方法（簡素化版）】
    その時点で直近に「確定している」決算期の年間DPS（= res[fy]["DPS"]、
    Step2の財務テーブルやExcelと同じ値）を、次の決算期が確定するまでの
    1年間そのまま固定値として使う（TTM DPSを1年間据え置き）。

    例：3月決算で fys=[2024, 2025, 2026]（2024.3期〜2026.3期が確定済み）の場合、
      2024.4.1 〜 2025.3.31 は 2024.3期の確定DPSを使用
      2025.4.1 〜 2026.3.31 は 2025.3期の確定DPSを使用
      2026.4.1 〜 直近営業日  は 2026.3期の確定DPSを使用
    グラフの表示期間は「fys内で最も古い決算期の翌日」〜「レポート作成時点の
    前営業日（週末なら金曜日）」。これにより、進行中の決算期でまだ配当が
    確定していなくても、直近の確定済みDPSを使って株価と同じ期間まで
    実績配当利回りを途切れさせずに描画できる。
    """
    if not fys:
        return None

    t = ticker if hasattr(ticker, "info") else yf.Ticker(ticker)

    # 表示開始日 = fys内で最も古い決算期の決算日の翌日
    oldest_fy = min(fys)
    fy_end_date = pd.Timestamp(year=oldest_fy, month=fiscal_month, day=1) + pd.offsets.MonthEnd(0)
    start = (fy_end_date + pd.Timedelta(days=1)).tz_localize(TZ_JP)

    # 表示終了日 = レポート作成時点の前営業日（週末なら金曜日）
    today = datetime.now(TZ_JP)
    as_of = previous_business_day(today)
    # yfinanceのend引数は指定日を含まない（exclusive）ため、+1日してas_ofを含める
    end = as_of + timedelta(days=1)

    if start >= end:
        return None

    hist = t.history(start=start, end=end)
    if hist.empty:
        return None

    def fy_of(dt):
        d = dt.tz_convert(TZ_JP) if dt.tzinfo else dt
        return d.year if d.month <= fiscal_month else d.year + 1

    def dps_for_date(dt):
        # その日付が属する決算期（未確定）の1つ前＝直近に確定している決算期のDPSを使う
        prior_fy = fy_of(dt) - 1
        return res.get(prior_fy, {}).get("DPS")

    hist["dps_ttm"] = [dps_for_date(idx) for idx in hist.index]
    hist["yield"] = hist["dps_ttm"] / hist["Close"] * 100

    # 株価ラインは期間内すべての営業日で途切れさせずに表示する。
    # 利回りラインは「対応する確定DPSが判明している区間」のみプロットする
    # （無配＝DPS=0の期間は0%として正しく表示され、データが無い区間のみ欠測にする）。
    yield_hist = hist.dropna(subset=["yield"])

    avg_y = yield_hist["yield"].mean() if not yield_hist.empty else None
    fore_y = forecast_dps / price * 100 if price else 0

    if avg_y is None or pd.isna(avg_y):
        verdict, vc = "—", COLORS["gray"]
    elif fore_y > avg_y * 1.1:
        verdict, vc = "割安", "#27AE60"
    elif fore_y < avg_y * 0.9:
        verdict, vc = "割高", "#E74C3C"
    else:
        verdict, vc = "中立", "#E67E22"

    fig, ax_p = plt.subplots(figsize=(9, 4.5))
    fig.patch.set_facecolor("#F2F2F2")
    ax_p.set_facecolor("#FFFFFF")

    ax_p.plot(hist.index, hist["Close"],
              label="株価（円）", color=COLORS["gray"], lw=1.5)
    ax_p.set_ylabel("株価（円）", color=COLORS["gray"], fontsize=10)
    ax_p.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"¥{v:,.0f}"))
    ax_p.tick_params(axis="y", labelcolor=COLORS["gray"])
    ax_p.xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m"))
    ax_p.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    ax_p.tick_params(axis="x", rotation=25)
    ax_p.grid(axis="both", ls="--", alpha=0.3)

    ax_y = ax_p.twinx()
    if not yield_hist.empty:
        ax_y.plot(yield_hist.index, yield_hist["yield"],
                  label="実績配当利回り（%）", color=COLORS["blue"], lw=1.5)
    if avg_y is not None and not pd.isna(avg_y):
        ax_y.axhline(avg_y, color="red", ls="--", lw=1.5,
                     label=f"過去平均: {avg_y:.2f}%")
    ax_y.plot(today, fore_y, marker="o", ms=11,
              color=vc, zorder=10, label=f"予想利回り: {fore_y:.2f}%（{verdict}）")
    ax_y.annotate(f"  {fore_y:.2f}%\n  {verdict}",
                  xy=(today, fore_y), fontsize=11,
                  fontweight="bold", color=vc, va="bottom")
    ax_y.set_ylabel("配当利回り（%）", color=COLORS["blue"], fontsize=10)
    ax_y.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.1f%%"))
    ax_y.tick_params(axis="y", labelcolor=COLORS["blue"])

    lines1, lbl1 = ax_p.get_legend_handles_labels()
    lines2, lbl2 = ax_y.get_legend_handles_labels()
    ax_y.legend(lines1 + lines2, lbl1 + lbl2, loc="upper left", fontsize=9)
    plt.title(f"{company}：配当利回りトレンドと格安判定", fontsize=11, pad=8)
    plt.tight_layout()
    return fig



# ══════════════════════════════════════════════════════════════
#  バッチ処理
# ══════════════════════════════════════════════════════════════

def read_batch_excel(file_bytes):
    """
    バッチ入力Excelを読み込み、行データのリストを返す。

    列定義（1行目ヘッダー）:
        コード, 決算月, 予想DPS,
        現預金-2, 現預金-1, 現預金,
        配当金支払総額-2, 配当金支払総額-1, 配当金支払総額

    空白行が来たら終了。
    戻り値: list of dict
    """
    import io
    from openpyxl import load_workbook as _lw

    wb = _lw(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        code = row[0]
        if code is None:
            break  # 空白行で終了
        try:
            rows.append({
                "code":         str(int(code)),
                "fiscal_month": int(row[1]),
                "forecast_dps": float(row[2]),
                "cash_vals":    [
                    float(row[3]) if row[3] is not None else None,  # -2期
                    float(row[4]) if row[4] is not None else None,  # -1期
                    float(row[5]) if row[5] is not None else None,  # 最新期
                ],
                "div_pay_vals": [
                    float(row[6]) if row[6] is not None else None,  # -2期
                    float(row[7]) if row[7] is not None else None,  # -1期
                    float(row[8]) if row[8] is not None else None,  # 最新期
                ],
            })
        except Exception as e:
            rows.append({"_parse_error": str(e), "_raw": row})
    return rows


def run_batch(file_bytes, progress_callback=None):
    """
    バッチExcelを読み込み、全銘柄のスクリーニングを実行する。

    progress_callback(current, total, code, name, status):
        進捗通知用コールバック（省略可）

    戻り値:
        {
            "succeeded": [{"code": ..., "name": ..., "history_id": ..., "filename": ...}, ...],
            "failed":    [{"code": ..., "reason": ...}, ...],
        }
    """
    import uuid as _uuid

    rows = read_batch_excel(file_bytes)
    total = len(rows)
    succeeded = []
    failed = []

    for i, row in enumerate(rows):
        # パースエラー行
        if "_parse_error" in row:
            failed.append({"code": str(row.get("_raw", ["?"])[0]), "reason": f"Excelの読み込みエラー: {row['_parse_error']}"})
            if progress_callback:
                progress_callback(i + 1, total, "?", "?", "failed")
            continue

        code         = row["code"]
        fiscal_month = row["fiscal_month"]
        forecast_dps = row["forecast_dps"]
        cash_vals    = row["cash_vals"]
        div_pay_vals = row["div_pay_vals"]

        try:
            # Step1: 基本情報取得
            basic = fetch_basic(code)
            name    = basic["name"]
            ticker  = basic["ticker"]
            price   = basic["price"]

            if progress_callback:
                progress_callback(i + 1, total, code, name, "fetching")

            # 予想利回り
            fore_yield = (forecast_dps / price * 100) if price else 0

            # Step2: 財務データ取得
            fys, res = fetch_financials(ticker, fiscal_month)

            years_label = [f"FY{fy}" for fy in fys]
            fcf_vals  = [res[fy].get("FCF") or 0 for fy in fys]
            debt_vals = [res[fy].get("有利子負債") or 0 for fy in fys]

            # FCF配当性向（app.pyのapi_step3と同じロジック）
            payout_vals = []
            for j, fy in enumerate(fys):
                fcf = res[fy].get("FCF")
                dv  = div_pay_vals[j] if j < len(div_pay_vals) else None
                if fcf is not None and dv is not None and fcf != 0:
                    payout_vals.append(round(dv / fcf * 100, 1))
                else:
                    payout_vals.append(None)

            div_pay_safe = [v or 0 for v in div_pay_vals]
            cash_safe    = [v or 0 for v in cash_vals]

            # Step3: グラフ・Excel生成
            fig_a = make_fig_a(name, years_label, fcf_vals, div_pay_safe, payout_vals)
            fig_b = make_fig_b(name, years_label, cash_safe, debt_vals, div_pay_safe)
            fig_c = make_fig_c(ticker, name, fiscal_month, forecast_dps, price, fys, res)

            out_path = write_excel(
                name, code, price, forecast_dps, fore_yield,
                fys, res, years_label,
                cash_vals, div_pay_vals,
                fig_a, fig_b, fig_c,
            )
            filename = os.path.basename(out_path)

            # 履歴に保存
            entry_id = _uuid.uuid4().hex[:12]
            archive_images = save_archive_images(entry_id, fig_a, fig_b, fig_c)
            history_entry = {
                "id":           entry_id,
                "code":         code,
                "ticker":       ticker,
                "name":         name,
                "created_at":   datetime.now(TZ_JP).isoformat(timespec="seconds"),
                "price":        price,
                "forecast_dps": forecast_dps,
                "fore_yield":   fore_yield,
                "fiscal_month": fiscal_month,
                "fys":          fys,
                "res":          {str(fy): v for fy, v in res.items()},
                "cash_vals":    cash_vals,
                "div_pay_vals": div_pay_vals,
                "filename":     filename,
                "images":       archive_images,
            }
            append_history_entry(history_entry)

            succeeded.append({
                "code":       code,
                "name":       name,
                "history_id": entry_id,
                "filename":   filename,
            })

            if progress_callback:
                progress_callback(i + 1, total, code, name, "success")

        except Exception as e:
            import traceback as _tb
            failed.append({"code": code, "reason": str(e)})
            if progress_callback:
                progress_callback(i + 1, total, code, "?", "failed")
        finally:
            for fig in [fig_a, fig_b, fig_c] if "fig_a" in dir() else []:
                try:
                    import matplotlib.pyplot as _plt
                    _plt.close(fig)
                except Exception:
                    pass

    return {"succeeded": succeeded, "failed": failed}


# ══════════════════════════════════════════════════════════════
#  Excel出力
# ══════════════════════════════════════════════════════════════

HDR_FILL = PatternFill("solid", fgColor="17375E")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
SUBHDR_FILL = PatternFill("solid", fgColor="B8CCE4")
SUBHDR_FONT = Font(bold=True, size=10)
CENT = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_cell(ws, row, col, val, font=None, fill=None, align=CENT, fmt=None, border=BORDER):
    c = ws.cell(row=row, column=col, value=val)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if fmt:
        c.number_format = fmt
    if border:
        c.border = border
    return c


def write_excel(company, code, price, forecast_dps, fore_yield,
                 fys, res, years_label,
                 cash_vals, div_pay_vals,
                 fig_a, fig_b, fig_c):
    """Excelブックを作成し OUTPUT_DIR に保存。保存先パスを返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "スクリーニング"

    for col, w in zip(range(1, 9), [22, 14, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A1:H1")
    set_cell(ws, 1, 1,
             f"{company}（{code}）　財務スクリーニング　{datetime.now().strftime('%Y/%m/%d')}",
             font=Font(color="FFFFFF", bold=True, size=12),
             fill=HDR_FILL)

    r = 3
    for label, val in [
        ("現在株価（円）", f"¥{price:,.0f}" if price else "—"),
        ("予想DPS（円）", f"¥{forecast_dps:,.0f}"),
        ("予想配当利回り", f"{fore_yield:.2f}%"),
    ]:
        set_cell(ws, r, 1, label, font=SUBHDR_FONT, fill=SUBHDR_FILL, align=CENT)
        set_cell(ws, r, 2, val, align=CENT)
        r += 1

    r = 7
    headers = ["指標"] + [f"FY{fy}" for fy in fys] + ["CAGR（3年）"]
    for ci, h in enumerate(headers, 1):
        set_cell(ws, r, ci, h, font=HDR_FONT, fill=HDR_FILL)
    r += 1

    for row in build_metric_rows(fys, res, cash_vals, div_pay_vals):
        label, vals, cg, fmt_str = row["label"], row["values"], row["cagr"], row["fmt"]
        set_cell(ws, r, 1, label, font=Font(bold=True), align=Alignment(horizontal="left"))
        for ci, v in enumerate(vals, 2):
            txt = fmt_str.format(v) if v is not None else "—"
            set_cell(ws, r, ci, txt, align=CENT)
        cg_txt = f"{cg:+.1f}%" if cg is not None else "—"
        set_cell(ws, r, 5, cg_txt,
                  font=Font(color="27AE60" if (cg or 0) >= 0 else "E74C3C", bold=True),
                  align=CENT)
        r += 1

    r += 1
    set_cell(ws, r, 1, "【短信入力】", font=SUBHDR_FONT, fill=SUBHDR_FILL)
    r += 1
    for label, vals in [("現預金（百万円）", cash_vals),
                          ("配当支払総額（百万円）", div_pay_vals)]:
        set_cell(ws, r, 1, label, font=Font(bold=True), align=Alignment(horizontal="left"))
        for ci, v in enumerate(vals, 2):
            txt = f"{v:,.1f}" if v is not None else "—"
            set_cell(ws, r, ci, txt, align=CENT)
        r += 1

    img_row = r + 2
    safe_name = safe_filename(company)
    p_a = os.path.join(OUTPUT_DIR, f"{safe_name}_tmp_a.png")
    p_b = os.path.join(OUTPUT_DIR, f"{safe_name}_tmp_b.png")
    p_c = os.path.join(OUTPUT_DIR, f"{safe_name}_tmp_c.png")

    for fig, path, anchor in [
        (fig_a, p_a, f"A{img_row}"),
        (fig_b, p_b, f"J{img_row}"),
        (fig_c, p_c, f"A{img_row + 26}"),
    ]:
        if fig is not None:
            fig.savefig(path, dpi=110, bbox_inches="tight")
            img = XLImage(path)
            img.width = 480
            img.height = 310
            ws.add_image(img, anchor)

    timestamp = datetime.now(TZ_JP).strftime("%Y%m%d_%H%M%S")
    out = os.path.join(OUTPUT_DIR, f"{safe_name}_{timestamp}_screening.xlsx")
    wb.save(out)

    for p in [p_a, p_b, p_c]:
        if p and os.path.exists(p):
            try:
                os.remove(p)
            except Exception:
                pass

    return out
